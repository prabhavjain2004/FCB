"""
TapNex superuser dashboard views for commission management and system analytics.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.models import User
from django.contrib.auth.forms import SetPasswordForm
from django.utils import timezone
from django.db import models
from datetime import datetime, timedelta, date
from decimal import Decimal
import json

from .models import TapNexSuperuser, CafeOwner
from .decorators import tapnex_superuser_required
from .commission_service import CommissionCalculator, RevenueTracker
from .forms import CommissionSettingsForm, CafeOwnerManagementForm
from booking.models import Booking, Game


@tapnex_superuser_required
def tapnex_dashboard(request):
    """TapNex superuser main dashboard with commission overview and analytics"""
    
    # Get or create TapNex superuser profile (NO defaults for commission/platform fee)
    tapnex_user, created = TapNexSuperuser.objects.get_or_create(
        user=request.user,
        defaults={
            'contact_email': request.user.email,
            # commission_rate and platform_fee must be set manually - no defaults
        }
    )
    
    # Get real-time metrics
    real_time_metrics = RevenueTracker.get_real_time_metrics()
    
    # Get revenue analytics for last 30 days
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    revenue_analytics = CommissionCalculator.get_revenue_analytics(start_date, end_date)
    
    # Get growth metrics
    growth_metrics = RevenueTracker.get_growth_metrics()
    
    # Get game revenue breakdown
    game_breakdown = CommissionCalculator.get_game_revenue_breakdown(start_date, end_date)
    
    # Get cafe owner information
    try:
        cafe_owner = CafeOwner.objects.select_related('user').first()
    except CafeOwner.DoesNotExist:
        cafe_owner = None
    
    # Recent bookings for monitoring
    recent_bookings = Booking.objects.filter(
        status__in=['CONFIRMED', 'IN_PROGRESS', 'COMPLETED', 'PENDING']
    ).select_related('customer__user', 'game', 'game_slot').order_by('-created_at')[:10]
    
    context = {
        'tapnex_user': tapnex_user,
        'real_time_metrics': real_time_metrics,
        'revenue_analytics': revenue_analytics,
        'growth_metrics': growth_metrics,
        'game_breakdown': game_breakdown,
        'cafe_owner': cafe_owner,
        'recent_bookings': recent_bookings,
        'user': request.user,
    }
    
    response = render(request, 'authentication/tapnex_dashboard.html', context)
    # Disable all caching for real-time updates
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@tapnex_superuser_required
def commission_settings(request):
    """Manage commission rates and platform fees"""
    
    tapnex_user = get_object_or_404(TapNexSuperuser, user=request.user)
    
    if request.method == 'POST':
        form = CommissionSettingsForm(request.POST, instance=tapnex_user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Commission settings updated successfully!')
            return redirect('authentication:tapnex_dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CommissionSettingsForm(instance=tapnex_user)
    
    # Calculate impact of new settings on recent bookings
    recent_bookings = Booking.objects.filter(
        status__in=['CONFIRMED', 'IN_PROGRESS', 'COMPLETED'],
        created_at__gte=timezone.now() - timedelta(days=7)
    )
    
    total_revenue = sum(booking.total_amount or Decimal('0.00') for booking in recent_bookings)
    current_commission = tapnex_user.calculate_commission(total_revenue)
    
    context = {
        'form': form,
        'tapnex_user': tapnex_user,
        'recent_revenue': total_revenue,
        'current_commission': current_commission,
        'recent_bookings_count': recent_bookings.count(),
    }
    
    response = render(request, 'authentication/commission_settings.html', context)
    # Disable all caching for real-time updates
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@tapnex_superuser_required
def revenue_reports(request):
    """TapNex Revenue Reports - Shows platform earnings (commission + platform fee)"""
    
    # Get date range from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'monthly')
    export_format = request.GET.get('export')
    
    # Default to current month if no dates provided
    if not start_date_str or not end_date_str:
        today = date.today()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Get TapNex-specific analytics (commission + platform fee)
    revenue_analytics = CommissionCalculator.get_tapnex_revenue_analytics(start_date, end_date)
    game_breakdown = CommissionCalculator.get_tapnex_game_revenue_breakdown(start_date, end_date)
    
    # Handle export requests
    if export_format == 'csv':
        return export_revenue_csv(revenue_analytics, game_breakdown, start_date, end_date)
    elif export_format == 'excel':
        return export_revenue_excel(revenue_analytics, game_breakdown, start_date, end_date)
    
    # Calculate monthly growth
    prev_month_start = start_date - timedelta(days=30)
    prev_month_end = start_date - timedelta(days=1)
    prev_analytics = CommissionCalculator.get_tapnex_revenue_analytics(prev_month_start, prev_month_end)
    
    current_revenue = revenue_analytics['totals']['tapnex_total_revenue']
    prev_revenue = prev_analytics['totals']['tapnex_total_revenue']
    
    if prev_revenue > 0:
        growth_rate = ((current_revenue - prev_revenue) / prev_revenue) * 100
    else:
        growth_rate = 100 if current_revenue > 0 else 0
    
    # Prepare chart data for frontend
    daily_chart_data = {
        'labels': [day['date'].strftime('%m/%d') for day in revenue_analytics['daily_trend']],
        'tapnex_revenue': [float(day['tapnex_revenue']) for day in revenue_analytics['daily_trend']],
        'commission': [float(day['commission']) for day in revenue_analytics['daily_trend']],
        'platform_fee': [float(day['platform_fee']) for day in revenue_analytics['daily_trend']],
        'bookings': [day['bookings'] for day in revenue_analytics['daily_trend']]
    }
    
    game_chart_data = {
        'labels': [game[0] for game in game_breakdown[:5]],  # Top 5 games
        'tapnex_revenue': [float(game[1]['tapnex_revenue']) for game in game_breakdown[:5]]
    }
    
    # Commission vs Platform Fee breakdown for pie chart
    commission_vs_fee_data = {
        'labels': ['Commission', 'Platform Fee'],
        'data': [
            float(revenue_analytics['totals']['total_commission']),
            float(revenue_analytics['totals']['total_platform_fee'])
        ]
    }
    
    context = {
        'revenue_analytics': revenue_analytics,
        'game_breakdown': game_breakdown,
        'growth_rate': growth_rate,
        'prev_revenue': prev_revenue,
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,
        'daily_chart_data': json.dumps(daily_chart_data),
        'game_chart_data': json.dumps(game_chart_data),
        'commission_vs_fee_data': json.dumps(commission_vs_fee_data),
    }
    
    response = render(request, 'authentication/revenue_reports.html', context)
    # Disable all caching for real-time updates
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@tapnex_superuser_required
def create_cafe_owner(request):
    """Create a new cafe owner account"""
    
    # Check if a cafe owner already exists
    if CafeOwner.objects.exists():
        messages.warning(request, 'A cafe owner account already exists. You can manage it instead.')
        return redirect('authentication:cafe_owner_management')
    
    if request.method == 'POST':
        from .forms import CafeOwnerRegistrationForm
        form = CafeOwnerRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cafe owner account created successfully!')
            return redirect('authentication:cafe_owner_management')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        from .forms import CafeOwnerRegistrationForm
        form = CafeOwnerRegistrationForm()
    
    context = {
        'form': form,
    }
    
    response = render(request, 'authentication/create_cafe_owner.html', context)
    # Disable all caching for real-time updates
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@tapnex_superuser_required
def cafe_owner_management(request):
    """Manage cafe owner account and settings"""
    
    try:
        cafe_owner = CafeOwner.objects.select_related('user').get()
    except CafeOwner.DoesNotExist:
        messages.error(request, 'No cafe owner account found. Please create one first.')
        return redirect('authentication:create_cafe_owner')
    
    if request.method == 'POST':
        form = CafeOwnerManagementForm(request.POST, instance=cafe_owner)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cafe owner account updated successfully!')
            return redirect('authentication:cafe_owner_management')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CafeOwnerManagementForm(instance=cafe_owner)
    
    # Get cafe owner's recent activity
    recent_games = Game.objects.filter(is_active=True).order_by('-created_at')[:5]
    recent_bookings = Booking.objects.filter(
        status__in=['CONFIRMED', 'IN_PROGRESS', 'COMPLETED']
    ).select_related('customer__user', 'game').order_by('-created_at')[:10]
    
    # Calculate cafe owner's revenue (net payout)
    from django.db import models
    total_revenue = Booking.objects.filter(
        status__in=['CONFIRMED', 'IN_PROGRESS', 'COMPLETED']
    ).aggregate(total=models.Sum('total_amount'))['total'] or Decimal('0.00')
    
    tapnex_user = TapNexSuperuser.objects.filter(user=request.user).first()
    if tapnex_user:
        commission_breakdown = tapnex_user.calculate_commission(total_revenue)
    else:
        commission_breakdown = {
            'gross_revenue': total_revenue,
            'net_payout': total_revenue
        }
    
    context = {
        'form': form,
        'cafe_owner': cafe_owner,
        'recent_games': recent_games,
        'recent_bookings': recent_bookings,
        'commission_breakdown': commission_breakdown,
    }
    
    response = render(request, 'authentication/cafe_owner_management.html', context)
    # Disable all caching for real-time updates
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@tapnex_superuser_required
@require_http_methods(["POST"])
def reset_cafe_owner_password(request):
    """Reset cafe owner password"""
    
    try:
        cafe_owner = CafeOwner.objects.select_related('user').get()
    except CafeOwner.DoesNotExist:
        messages.error(request, 'Cafe owner not found.')
        return redirect('authentication:cafe_owner_management')
    
    new_password = request.POST.get('new_password')
    confirm_password = request.POST.get('confirm_password')
    
    if not new_password or not confirm_password:
        messages.error(request, 'Both password fields are required.')
        return redirect('authentication:cafe_owner_management')
    
    if new_password != confirm_password:
        messages.error(request, 'Passwords do not match.')
        return redirect('authentication:cafe_owner_management')
    
    if len(new_password) < 8:
        messages.error(request, 'Password must be at least 8 characters long.')
        return redirect('authentication:cafe_owner_management')
    
    # Set new password
    cafe_owner.user.set_password(new_password)
    cafe_owner.user.save()
    
    messages.success(request, f'Password reset successfully for {cafe_owner.user.username}.')
    return redirect('authentication:cafe_owner_management')


@tapnex_superuser_required
def system_analytics(request):
    """System-wide analytics and monitoring"""
    
    # Get system metrics
    total_users = User.objects.count()
    total_customers = User.objects.filter(customer_profile__isnull=False).count()
    total_games = Game.objects.filter(is_active=True).count()
    
    # Booking statistics
    total_bookings = Booking.objects.count()
    confirmed_bookings = Booking.objects.filter(
        status__in=['CONFIRMED', 'IN_PROGRESS', 'COMPLETED']
    ).count()
    
    # Revenue trends (last 12 months)
    monthly_trends = []
    today = date.today()
    
    for i in range(12):
        # Calculate month
        if today.month - i <= 0:
            month = 12 + (today.month - i)
            year = today.year - 1
        else:
            month = today.month - i
            year = today.year
        
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        month_revenue = Booking.objects.filter(
            status__in=['CONFIRMED', 'IN_PROGRESS', 'COMPLETED'],
            created_at__date__gte=month_start,
            created_at__date__lte=month_end
        ).aggregate(total=models.Sum('total_amount'))['total'] or Decimal('0.00')
        
        monthly_trends.append({
            'month': month_start.strftime('%b %Y'),
            'revenue': month_revenue
        })
    
    monthly_trends.reverse()  # Chronological order
    
    # Calculate growth percentages
    for i in range(len(monthly_trends)):
        if i > 0:
            prev_revenue = monthly_trends[i-1]['revenue']
            curr_revenue = monthly_trends[i]['revenue']
            if prev_revenue > 0:
                growth = ((curr_revenue - prev_revenue) / prev_revenue) * 100
                monthly_trends[i]['growth'] = float(growth)
            else:
                monthly_trends[i]['growth'] = 0 if curr_revenue == 0 else 100
        else:
            monthly_trends[i]['growth'] = None
    
    # Booking type distribution
    booking_type_stats = {
        'private': Booking.objects.filter(booking_type='PRIVATE').count(),
        'shared': Booking.objects.filter(booking_type='SHARED').count()
    }
    
    # Peak hours analysis (last 30 days)
    from django.db.models import Count
    from django.db.models.functions import Extract
    
    peak_hours = Booking.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30),
        game_slot__isnull=False
    ).annotate(
        hour=Extract('game_slot__start_time', 'hour')
    ).values('hour').annotate(
        booking_count=Count('id')
    ).order_by('hour')
    
    context = {
        'system_stats': {
            'total_users': total_users,
            'total_customers': total_customers,
            'total_games': total_games,
            'total_bookings': total_bookings,
            'confirmed_bookings': confirmed_bookings,
            'conversion_rate': (confirmed_bookings / total_bookings * 100) if total_bookings > 0 else 0
        },
        'monthly_trends': monthly_trends,
        'booking_type_stats': booking_type_stats,
        'peak_hours': list(peak_hours),
        'monthly_chart_data': json.dumps({
            'labels': [trend['month'] for trend in monthly_trends],
            'revenue': [float(trend['revenue']) for trend in monthly_trends]
        }),
        'booking_type_chart_data': json.dumps({
            'labels': ['Private Bookings', 'Shared Bookings'],
            'data': [booking_type_stats['private'], booking_type_stats['shared']]
        })
    }
    
    response = render(request, 'authentication/system_analytics.html', context)
    # Disable all caching for real-time updates
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@tapnex_superuser_required
def ajax_revenue_data(request):
    """AJAX endpoint for real-time revenue data updates"""
    
    # Get real-time metrics
    metrics = RevenueTracker.get_real_time_metrics()
    
    # Format for JSON response
    response_data = {
        'today_revenue': float(metrics['today']['revenue']),
        'today_bookings': metrics['today']['bookings'],
        'today_commission': float(metrics['today']['commission']),
        'month_revenue': float(metrics['month']['revenue']),
        'month_bookings': metrics['month']['bookings'],
        'active_bookings': metrics['active']['in_progress'],
        'pending_payments': metrics['active']['pending_payments'],
        'timestamp': timezone.now().isoformat()
    }
    
    return JsonResponse(response_data)



def export_revenue_csv(revenue_analytics, game_breakdown, start_date, end_date):
    """Export revenue report as CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="tapnex_revenue_{start_date}_to_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow(['TapNex Revenue Report'])
    writer.writerow([f'Period: {start_date} to {end_date}'])
    writer.writerow([])
    
    # Summary
    writer.writerow(['Summary'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Total TapNex Revenue', f"₹{revenue_analytics['totals']['tapnex_total_revenue']}"])
    writer.writerow(['Total Commission', f"₹{revenue_analytics['totals']['total_commission']}"])
    writer.writerow(['Total Platform Fee', f"₹{revenue_analytics['totals']['total_platform_fee']}"])
    writer.writerow(['Total Bookings', revenue_analytics['totals']['total_bookings']])
    writer.writerow(['Average Revenue per Booking', f"₹{revenue_analytics['totals']['avg_revenue_per_booking']}"])
    writer.writerow([])
    
    # Game Breakdown
    writer.writerow(['Game Performance'])
    writer.writerow(['Game', 'Bookings', 'TapNex Revenue', 'Commission', 'Platform Fee', 'Private Bookings', 'Shared Bookings'])
    for game_name, stats in game_breakdown:
        writer.writerow([
            game_name,
            stats['bookings'],
            f"₹{stats['tapnex_revenue']}",
            f"₹{stats['commission']}",
            f"₹{stats['platform_fee']}",
            stats['private_bookings'],
            stats['shared_bookings']
        ])
    writer.writerow([])
    
    # Daily Trend
    writer.writerow(['Daily Revenue Trend'])
    writer.writerow(['Date', 'TapNex Revenue', 'Commission', 'Platform Fee', 'Bookings'])
    for day in revenue_analytics['daily_trend']:
        writer.writerow([
            day['date'],
            f"₹{day['tapnex_revenue']}",
            f"₹{day['commission']}",
            f"₹{day['platform_fee']}",
            day['bookings']
        ])
    
    return response


def export_revenue_excel(revenue_analytics, game_breakdown, start_date, end_date):
    """Export revenue report as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from io import BytesIO
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        messages.warning(request, 'Excel export requires openpyxl. Downloading as CSV instead.')
        return export_revenue_csv(revenue_analytics, game_breakdown, start_date, end_date)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TapNex Revenue"
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    
    # Header
    ws['A1'] = 'TapNex Revenue Report'
    ws['A1'].font = header_font
    ws['A2'] = f'Period: {start_date} to {end_date}'
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    
    # Summary
    row = 4
    ws[f'A{row}'] = 'Summary'
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    ws[f'A{row}'] = 'Metric'
    ws[f'B{row}'] = 'Value'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    summary_data = [
        ['Total TapNex Revenue', f"₹{revenue_analytics['totals']['tapnex_total_revenue']}"],
        ['Total Commission', f"₹{revenue_analytics['totals']['total_commission']}"],
        ['Total Platform Fee', f"₹{revenue_analytics['totals']['total_platform_fee']}"],
        ['Total Bookings', revenue_analytics['totals']['total_bookings']],
        ['Average Revenue per Booking', f"₹{revenue_analytics['totals']['avg_revenue_per_booking']}"],
    ]
    
    for metric, value in summary_data:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        row += 1
    
    # Game Breakdown
    row += 2
    ws[f'A{row}'] = 'Game Performance'
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    headers = ['Game', 'Bookings', 'TapNex Revenue', 'Commission', 'Platform Fee', 'Private', 'Shared']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
    row += 1
    
    for game_name, stats in game_breakdown:
        ws[f'A{row}'] = game_name
        ws[f'B{row}'] = stats['bookings']
        ws[f'C{row}'] = f"₹{stats['tapnex_revenue']}"
        ws[f'D{row}'] = f"₹{stats['commission']}"
        ws[f'E{row}'] = f"₹{stats['platform_fee']}"
        ws[f'F{row}'] = stats['private_bookings']
        ws[f'G{row}'] = stats['shared_bookings']
        row += 1
    
    # Daily Trend
    row += 2
    ws[f'A{row}'] = 'Daily Revenue Trend'
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    headers = ['Date', 'TapNex Revenue', 'Commission', 'Platform Fee', 'Bookings']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
    row += 1
    
    for day in revenue_analytics['daily_trend']:
        ws[f'A{row}'] = str(day['date'])
        ws[f'B{row}'] = f"₹{day['tapnex_revenue']}"
        ws[f'C{row}'] = f"₹{day['commission']}"
        ws[f'D{row}'] = f"₹{day['platform_fee']}"
        ws[f'E{row}'] = day['bookings']
        row += 1
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tapnex_revenue_{start_date}_to_{end_date}.xlsx"'
    
    return response
